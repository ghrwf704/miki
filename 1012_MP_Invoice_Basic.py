import subprocess
import openpyxl as excel
import re
import datetime
import os
import csv
import requests
import time
import json
from PIL import Image
import shutil
import numpy as np
import pdf2image
import cv2 as cv
import math
import codecs

####################使い方###########################
# 本スクリプト（または実行）ファイルと自動転記用のエクセル、設定ファイル 
# 1010_OCR_Regist.csv を同一フォルダにおく。
# 読取りファイル(PDF 形式)は、「Data」フォルダーに格納、
#（実行するときはＰＤＦファイル以外は削除しておく）
# 実行結果は、 定期支払チェックリスト.ｘｌｓｘ に保存される。
###################################################

def readImg(temp,secret,apiurl):
    #アップロード用のphpファイルに接続
    url = 'http://mikipulley.sakura.ne.jp/microsoft/azuru/ocr/img/upload.php'
    file = {'upfile': open(os.getcwd()+"\\Data\\"+temp, 'rb')}
    #res = requests.post(url, files = file)#画像ファイルなどをアップするのに必要
    res = requests.post(url, files = file, proxies = proxies) # proxy設定時
    #ocr用のコマンドを作成
    url = apiurl
    payload = "{\r\n    \"images\": [\r\n      {\r\n        \"format\": \"png\",\r\n        \"name\": \"All\",\r\n        \"data\": null,\r\n        \"url\": \""+ 'http://mikipulley.sakura.ne.jp/microsoft/azuru/ocr/img/files/'+temp+"\"\r\n      }\r\n    ],\r\n    \"lang\": \"ja\",\r\n    \"requestId\": \"string\",\r\n    \"resultType\": \"string\",\r\n    \"timestamp\": 1582166299,\r\n    \"version\": \"V1\"\r\n}"
    headers = {
            'Content-Type': 'application/json',
            'X-OCR-SECRET': secret
            }
    try:
        #ocrの結果を保存
        #response = requests.request("POST", url, headers=headers, data = payload)
        response = requests.request("POST", url, headers=headers, data = payload, proxies = proxies)  #proxy設定時
        #戻り値にエラーがあった場合はすべての値を空白にして返す（合致するテンプレートがない場合もエラー）
        f = codecs.open("Data\\"+temp.replace(".png",".txt"), 'w', 'utf-8')
        json.dump(response.json(), f, ensure_ascii=False)
        f.close()
        #保存しておいたテキストからjsonとして読み込む
        json_open = codecs.open('Data\\'+tifs[i].replace(".png",".txt"), 'r', 'utf-8')
        response = json.load(json_open)
        json_open.close()
        company=response["images"][0]["title"]["inferText"].replace("(","").replace(")","").replace(":","").replace(",","").replace(".","").replace("　","").replace(" ","").replace("\n","")
        iNo=response["images"][0]["fields"][0]["inferText"].replace("(","").replace(")","").replace(":","").replace(",","").replace(".","").replace("　","").replace(" ","").replace("\n","")
        tAmount=response["images"][0]["fields"][1]["inferText"].replace("(","").replace(")","").replace(":","").replace("　","").replace("\n","")
    except:
        company=""
        iNo=""
        tAmount=""
    return company,iNo,tAmount

def getCompanyInfo(company):
    temp=["Kari"]
    with open('1012_OCR_Regist.csv') as f:
        rows=csv.reader(f)
        for temp in rows:
            if ""==company:
                break
            if company in str(temp):
                return temp[0:16]
            if len(temp)!=16:
                break
    return ["","","","","","","","","","","","","","","",""]


import socks
# proxy設定時
proxies = {
  'https': 'swg-proxy.mikipulley.co.jp:8080',
}
socks.set_default_proxy(socks.SOCKS5, "swg-proxy.mikipulley.co.jp", 8080)

#2020/11/18 追加  Start
#一時的にpathに「poppler\\bin」を追加
#cwd = os.getcwd()
#bin_path = os.path.join(cwd, 'poppler\\bin')
#os.environ['PATH'] = '{};{}'.format(bin_path, os.environ['PATH'])
#q = os.chdir("Data")
#2020/11/18 追加   End

#「poppler\\bin」無い場合
os.getcwd()                     
q = os.chdir("Data")            
#「poppler\\bin」無い場合


filepath = os.listdir(q)
#  os.system("del /q \""+os.getcwd().replace("\\Data","\\CVImage")+"\\*.*\"")
for file in filepath:
    basename = os.path.basename(file)
    pdfimages = pdf2image.convert_from_path(file)    
    cvimage = np.asarray(pdfimages[0])
    cvimage = cv.cvtColor(cvimage, cv.COLOR_RGB2BGR)
##  cvimage = cv.resize(cvimage, (960, 640))
    temp=str(basename).replace('.pdf','.png')
    cv.imwrite(temp, cvimage)
    cv.waitKey(-1) 

#カレントディレクトリー変更
path=os.getcwd()
files = os.listdir(path)
files_file = [f for f in files if os.path.isfile(os.path.join(path, f))]

filepath = os.listdir(q)

#ファイル名から" "空白と"#"を置換
for file in filepath:
    os.rename(file,file.replace(" ","_").replace("#","_"))

os.chdir("../")

#エクセル起動
wb = excel.Workbook()
wb['Sheet'].cell(row=1, column=1,value='伝票№')
wb['Sheet'].cell(row=1, column=2,value='固定値"RAK"')
wb['Sheet'].cell(row=1, column=3,value='会社ｺｰﾄﾞ "E0"')    #テスト  E0
wb['Sheet'].cell(row=1, column=4,value='固定値"002845"')
wb['Sheet'].cell(row=1, column=5,value='固定値"1016"')
wb['Sheet'].cell(row=1, column=6,value='固定値"002845"')
wb['Sheet'].cell(row=1, column=7,value='仕訳日')
wb['Sheet'].cell(row=1, column=8,value='固定値"0"')
wb['Sheet'].cell(row=1, column=9,value='固定値"00"')
wb['Sheet'].cell(row=1, column=10,value='仕訳日')
wb['Sheet'].cell(row=1, column=11,value='空白')
wb['Sheet'].cell(row=1, column=12,value='固定値"1"')
wb['Sheet'].cell(row=1, column=13,value='固定値"0"')
wb['Sheet'].cell(row=1, column=14,value='空白')
wb['Sheet'].cell(row=1, column=15,value='空白')
wb['Sheet'].cell(row=1, column=16,value='空白')
wb['Sheet'].cell(row=1, column=17,value='空白')
wb['Sheet'].cell(row=1, column=18,value='空白')
wb['Sheet'].cell(row=1, column=19,value='空白')
wb['Sheet'].cell(row=1, column=20,value='空白')
wb['Sheet'].cell(row=1, column=21,value='空白')
wb['Sheet'].cell(row=1, column=22,value='空白')
wb['Sheet'].cell(row=1, column=23,value='空白')
wb['Sheet'].cell(row=1, column=24,value='空白')
wb['Sheet'].cell(row=1, column=25,value='空白')
wb['Sheet'].cell(row=1, column=26,value='空白')
wb['Sheet'].cell(row=1, column=27,value='空白')
wb['Sheet'].cell(row=1, column=28,value='空白')
wb['Sheet'].cell(row=1, column=29,value='空白')
wb['Sheet'].cell(row=1, column=30,value='空白')
wb['Sheet'].cell(row=1, column=31,value='空白')
wb['Sheet'].cell(row=1, column=32,value='空白')
wb['Sheet'].cell(row=1, column=33,value='空白')
wb['Sheet'].cell(row=1, column=34,value='空白')
wb['Sheet'].cell(row=1, column=35,value='空白')
wb['Sheet'].cell(row=1, column=36,value='摘要')
wb['Sheet'].cell(row=1, column=37,value='空白')
wb['Sheet'].cell(row=1, column=38,value='空白')
wb['Sheet'].cell(row=1, column=39,value='空白')
wb['Sheet'].cell(row=1, column=40,value='伝票内連番')
wb['Sheet'].cell(row=1, column=41,value='貸借区分（借方ﾏｲﾅｽの場合0→1、貸方ﾏｲﾅｽの場合1→0）')
wb['Sheet'].cell(row=1, column=42,value='勘定科目ｺｰﾄﾞ')
wb['Sheet'].cell(row=1, column=43,value='負担部門ｺｰﾄﾞ')
wb['Sheet'].cell(row=1, column=44,value='空白')
wb['Sheet'].cell(row=1, column=45,value='細目ｺｰﾄﾞ')
wb['Sheet'].cell(row=1, column=46,value='空白')
wb['Sheet'].cell(row=1, column=47,value='空白')
wb['Sheet'].cell(row=1, column=48,value='空白')
wb['Sheet'].cell(row=1, column=49,value='集計拡張ｺｰﾄﾞ1（旧補助元帳）')
wb['Sheet'].cell(row=1, column=50,value='空白')
wb['Sheet'].cell(row=1, column=51,value='集計拡張ｺｰﾄﾞ2（従業員№）')
wb['Sheet'].cell(row=1, column=52,value='空白')
wb['Sheet'].cell(row=1, column=53,value='稟議№')
wb['Sheet'].cell(row=1, column=54,value='空白')
wb['Sheet'].cell(row=1, column=55,value='空白')
wb['Sheet'].cell(row=1, column=56,value='空白')
wb['Sheet'].cell(row=1, column=57,value='空白')
wb['Sheet'].cell(row=1, column=58,value='空白')
wb['Sheet'].cell(row=1, column=59,value='空白')
wb['Sheet'].cell(row=1, column=60,value='空白')
wb['Sheet'].cell(row=1, column=61,value='空白')
wb['Sheet'].cell(row=1, column=62,value='空白')
wb['Sheet'].cell(row=1, column=63,value='空白')
wb['Sheet'].cell(row=1, column=64,value='空白')
wb['Sheet'].cell(row=1, column=65,value='空白')
wb['Sheet'].cell(row=1, column=66,value='空白')
wb['Sheet'].cell(row=1, column=67,value='空白')
wb['Sheet'].cell(row=1, column=68,value='取引先ｺｰﾄﾞ')
wb['Sheet'].cell(row=1, column=69,value='空白')
wb['Sheet'].cell(row=1, column=70,value='空白')
wb['Sheet'].cell(row=1, column=71,value='請求支払先ｺｰﾄﾞ')
wb['Sheet'].cell(row=1, column=72,value='空白')
wb['Sheet'].cell(row=1, column=73,value='空白')
wb['Sheet'].cell(row=1, column=74,value='空白')
wb['Sheet'].cell(row=1, column=75,value='空白')
wb['Sheet'].cell(row=1, column=76,value='空白')
wb['Sheet'].cell(row=1, column=77,value='空白')
wb['Sheet'].cell(row=1, column=78,value='空白')
wb['Sheet'].cell(row=1, column=79,value='空白')
wb['Sheet'].cell(row=1, column=80,value='空白')
wb['Sheet'].cell(row=1, column=81,value='空白')
wb['Sheet'].cell(row=1, column=82,value='空白')
wb['Sheet'].cell(row=1, column=83,value='空白')
wb['Sheet'].cell(row=1, column=84,value='空白')
wb['Sheet'].cell(row=1, column=85,value='空白')
wb['Sheet'].cell(row=1, column=86,value='空白')
wb['Sheet'].cell(row=1, column=87,value='空白')
wb['Sheet'].cell(row=1, column=88,value='消費税区分ｺｰﾄﾞ')
wb['Sheet'].cell(row=1, column=89,value='空白')
wb['Sheet'].cell(row=1, column=90,value='消費税率区分')
wb['Sheet'].cell(row=1, column=91,value='金額（借方ﾏｲﾅｽの場合ﾌﾟﾗｽ化、貸方ﾏｲﾅｽの場合ﾌﾟﾗｽ化）')
wb['Sheet'].cell(row=1, column=92,value='空白')
wb['Sheet'].cell(row=1, column=93,value='参考消費税額（ﾏｲﾅｽの場合は絶対値）')
wb['Sheet'].cell(row=1, column=94,value='空白')
wb['Sheet'].cell(row=1, column=95,value='消費税区分')
wb['Sheet'].cell(row=1, column=96,value='空白')
wb['Sheet'].cell(row=1, column=97,value='空白')
wb['Sheet'].cell(row=1, column=98,value='空白')
wb['Sheet'].cell(row=1, column=99,value='空白')
wb['Sheet'].cell(row=1, column=100,value='空白')
wb['Sheet'].cell(row=1, column=101,value='空白')
wb['Sheet'].cell(row=1, column=102,value='空白')
wb['Sheet'].cell(row=1, column=103,value='空白')
wb['Sheet'].cell(row=1, column=104,value='空白')
wb['Sheet'].cell(row=1, column=105,value='空白')
wb['Sheet'].cell(row=1, column=106,value='空白')
wb['Sheet'].cell(row=1, column=107,value='空白')
wb['Sheet'].cell(row=1, column=108,value='空白')
wb['Sheet'].cell(row=1, column=109,value='空白')
wb['Sheet'].cell(row=1, column=110,value='空白')
wb['Sheet'].cell(row=1, column=111,value='空白')
wb['Sheet'].cell(row=1, column=112,value='空白')
wb['Sheet'].cell(row=1, column=113,value='空白')
wb['Sheet'].cell(row=1, column=114,value='空白')
wb['Sheet'].cell(row=1, column=115,value='空白')
wb['Sheet'].cell(row=1, column=116,value='空白')
wb['Sheet'].cell(row=1, column=117,value='空白')
wb['Sheet'].cell(row=1, column=118,value='空白')
wb['Sheet'].cell(row=1, column=119,value='空白')
wb['Sheet'].cell(row=1, column=120,value='空白')
wb['Sheet'].cell(row=1, column=121,value='空白')
wb['Sheet'].cell(row=1, column=122,value='空白')
wb['Sheet'].cell(row=1, column=123,value='空白')
wb['Sheet'].cell(row=1, column=124,value='空白')
wb['Sheet'].cell(row=1, column=125,value='空白')
wb['Sheet'].cell(row=1, column=126,value='空白')
wb['Sheet'].cell(row=1, column=127,value='空白')
wb['Sheet'].cell(row=1, column=128,value='空白')
wb['Sheet'].cell(row=1, column=129,value='空白')
wb['Sheet'].cell(row=1, column=130,value='空白')
wb['Sheet'].cell(row=1, column=131,value='空白')
wb['Sheet'].cell(row=1, column=132,value='空白')
wb['Sheet'].cell(row=1, column=133,value='空白')
wb['Sheet'].cell(row=1, column=134,value='空白')
wb['Sheet'].cell(row=1, column=135,value='空白')
wb['Sheet'].cell(row=1, column=136,value='空白')
wb['Sheet'].cell(row=1, column=137,value='空白')
wb['Sheet'].cell(row=1, column=138,value='空白')
wb['Sheet'].cell(row=1, column=139,value='空白')
wb['Sheet'].cell(row=1, column=140,value='空白')
wb['Sheet'].cell(row=1, column=141,value='空白')
wb['Sheet'].cell(row=1, column=142,value='空白')
wb['Sheet'].cell(row=1, column=143,value='空白')
wb['Sheet'].cell(row=1, column=144,value='空白')
wb['Sheet'].cell(row=1, column=145,value='空白')
wb['Sheet'].cell(row=1, column=146,value='空白')
wb['Sheet'].cell(row=1, column=147,value='空白')
wb['Sheet'].cell(row=1, column=148,value='空白')
wb['Sheet'].cell(row=1, column=149,value='空白')
wb['Sheet'].cell(row=1, column=150,value='空白')
wb['Sheet'].cell(row=1, column=151,value='空白')
wb['Sheet'].cell(row=1, column=152,value='行摘要')
wb['Sheet'].cell(row=1, column=153,value='空白')
wb['Sheet'].cell(row=1, column=154,value='空白')
wb['Sheet'].cell(row=1, column=155,value='空白')
wb['Sheet'].cell(row=1, column=156,value='空白')
wb['Sheet'].cell(row=1, column=157,value='契約番号')
wb['Sheet'].cell(row=1, column=158,value='空白')
wb['Sheet'].cell(row=1, column=159,value='空白')
wb['Sheet'].cell(row=1, column=160,value='空白')
wb['Sheet'].cell(row=1, column=161,value='更新ｻﾌﾞｼｽﾃﾑ区分')
wb['Sheet'].cell(row=1, column=162,value='空白')
wb['Sheet'].cell(row=1, column=163,value='空白')
wb['Sheet'].cell(row=1, column=164,value='空白')
wb['Sheet'].cell(row=1, column=165,value='空白')
wb['Sheet'].cell(row=1, column=166,value='空白')
wb['Sheet'].cell(row=1, column=167,value='空白')
wb['Sheet'].cell(row=1, column=168,value='空白')
wb['Sheet'].cell(row=1, column=169,value='空白')
wb['Sheet'].cell(row=1, column=170,value='空白')
wb['Sheet'].cell(row=1, column=171,value='空白')
wb['Sheet'].cell(row=1, column=172,value='空白')
wb['Sheet'].cell(row=1, column=173,value='空白')
wb['Sheet'].cell(row=1, column=174,value='空白')
wb['Sheet'].cell(row=1, column=175,value='空白')
wb['Sheet'].cell(row=1, column=176,value='空白')
wb['Sheet'].cell(row=1, column=177,value='空白')
wb['Sheet'].cell(row=1, column=178,value='空白')
wb['Sheet'].cell(row=1, column=179,value='空白')
wb['Sheet'].cell(row=1, column=180,value='空白')
wb['Sheet'].cell(row=1, column=181,value='空白')
wb['Sheet'].cell(row=1, column=182,value='空白')
wb['Sheet'].cell(row=1, column=183,value='空白')
wb['Sheet'].cell(row=1, column=184,value='空白')
wb['Sheet'].cell(row=1, column=185,value='空白')
wb['Sheet'].cell(row=1, column=186,value='空白')
wb['Sheet'].cell(row=1, column=187,value='空白')
wb['Sheet'].cell(row=1, column=188,value='空白')

wb1 = excel.Workbook()
wb1['Sheet'].cell(row=1, column=1,value='UP伝票№')
wb1['Sheet'].cell(row=1, column=2,value='仕訳日')
wb1['Sheet'].cell(row=1, column=3,value='取引先ｺｰﾄﾞ')
wb1['Sheet'].cell(row=1, column=4,value='摘要')
wb1['Sheet'].cell(row=1, column=5,value='請求書No')
wb1['Sheet'].cell(row=1, column=6,value='金額')
wb1['Sheet'].cell(row=1, column=7,value='消費税')
wb1['Sheet'].cell(row=1, column=8,value='税フラグ')
wb1['Sheet'].cell(row=1, column=9,value='RGT_伝票№')
wb1['Sheet'].cell(row=1, column=10,value='RGT_Unique')
wb1['Sheet'].cell(row=1, column=11,value='PDFファイル名')


#現在の年月日情報を取得
time_now = datetime.datetime.now()
year = time_now.strftime("%Y")
month = time_now.strftime("%m")
day = time_now.strftime("%d")

#対象のファイル名 .txt を書き出す（ｓｔｒ１にカンマ区切りで格納しておく)
subprocess.call("dir /B "+os.getcwd()+"\\Data\\*.png>"+os.getcwd()+"\\tempfile.txt",shell=True)
str1=""
t=open('tempfile.txt')
f=t.readline()
while f:
	if ""==f:
		continue
	if "tmp.png"==f:
		continue
	str1=str1+f+","
	f=t.readline()
t.close()
str1=str1.replace("\n","")

#ファイル名が格納された配列作成（tifs配列にはファイル名のみが入っている）
tifs=str1.split(",")
os.remove("tempfile.txt")
ocr_output=""
ii=0
company="Kari"
jj=201                                                      #UPLoadファイル用連番 ← 伝票№から変更 2020/11/30

for i in range(0, len(tifs)):
    try:
        if tifs[i]=="":
            break
        #画像、テンプレートが一致したら、INVOICE番号と、合計金額を取得
        [company,iNo,tAmount]=readImg(tifs[i],'QnBRWWdLeU5Yckx1d2JFUFRyVXRaaldvbldjRlNDbEU=','https://x8g7u5g3e0.apigw.ntruss.com/custom/v1/370/586809a84ad75b52815a003a1b14e7b7a5128bb3c035b70924e57720ace4f130/infer')
        #if company=="" or tAmount=="":
        #    [company,iNo,tAmount]=readImg(tifs[i],'d1NIc0Rha21iSWRmWGhrcXhkeVVxa2hzaWhrYUpaTmQ=',"https://x8g7u5g3e0.apigw.ntruss.com/custom/v1/142/6a6b3f6f073a097b717ebc912bc844c01612087482d602f094292fef8e3ec9f1/infer")
        if company=="" or tAmount=="":
            with open("1012_ErrorList_"+str(datetime.date.today().strftime('%Y%m%d'))+".txt","a+") as f:
                f.write(str([tifs[i],company+",InvoiceNo:"+str(iNo),"Amount:"+str(tAmount)])+"\n")
            continue
        #companyのデータをresist.csvから取得
        #csvのA列のデータ==companyになってるのを探して行情報を返す
        [Unique,伝票,摘要,勘定科目コード,負担部門コード,細目コード,集計拡張コード,稟議,取引先コード,消費税区分コード,
         消費税率区分,消費税区分,税フラグ,税率,SampleName,登録ファイル名]=getCompanyInfo(company)
        tAmount=tAmount.split("\n")[-1]
        #tAmountに格納されている文字列データのうち、最下段文字列の1-0のみmを取得
        #改行以降を削除したうえで、数値のみを連結させる
        match = re.findall(r'[0-9]+', re.sub('\n.+',"",tAmount))
        s=""
        #連結（飛び飛びでもかいつまんで連結させる）※1,2,er45,t77→124577→csvファイルの一番右の列がで割る（124577 or 1245.77）
        for m in match:
    	    s=s+str(m)
        if s=="":
            s=0
        tAmount=int(s)
        #iNoを正規表現で抽出する 空白削除
        if iNo==0 or iNo=="0":
            iNo=""
        if company=="" or tAmount==0:#(conf1)<0.8 or (conf2)<0.8
            with open("1012_ErrorList_"+str(datetime.date.today().strftime('%Y%m%d'))+".txt","a+") as f:
                f.write(str([tifs[i],company+",InvoiceNo:"+str(iNo),"Amount:"+str(tAmount)])+"\n")
            continue
        #excel操作
        if int(day)<=23:                                         #処理日をいつまでするか？
            仕入日=str(year)+str(month)+"20"
        else:
            if int(month)==12:
                month==0
            仕入日=str(year)+str(int(month)+1)+"20"     
        if int(税フラグ)==1:        
            消費税=round(tAmount/11)
        else:
            tAmount=math.floor(tAmount*1.1)
            消費税=round(tAmount/11)
        #チェックリストに記入
        wb1['Sheet'].cell(row=ii+2, column=1,value=jj)          #UP連番 ← 伝票№から変更 2020/11/30
        wb1['Sheet'].cell(row=ii+2, column=2,value=仕入日)
        wb1['Sheet'].cell(row=ii+2, column=3,value=取引先コード)
        wb1['Sheet'].cell(row=ii+2, column=4,value=摘要)
        wb1['Sheet'].cell(row=ii+2, column=5,value=iNo)
        wb1['Sheet'].cell(row=ii+2, column=6,value=tAmount)
        wb1['Sheet'].cell(row=ii+2, column=7,value=消費税)
        wb1['Sheet'].cell(row=ii+2, column=8,value=税フラグ)
        wb1['Sheet'].cell(row=ii+2, column=9,value=伝票)
        wb1['Sheet'].cell(row=ii+2, column=10,value=Unique)
        wb1['Sheet'].cell(row=ii+2, column=11,value=tifs[i].replace(".png",".pdf"))
        wb1.save('1012_Teiki_CheckList.xlsx')
        #upLoadするファイルに記入
        #読み取るたびに2列ずつ書き込む
        wb['Sheet'].cell(row=(2*ii)+2, column=1,value=jj)        #UP連番 ← 伝票№から変更 2020/11/30
        wb['Sheet'].cell(row=(2*ii)+3, column=1,value=jj)        #UP連番 ← 伝票№から変更 2020/11/30
        wb['Sheet'].cell(row=(2*ii)+2, column=2,value="RAK")
        wb['Sheet'].cell(row=(2*ii)+3, column=2,value="RAK")
        wb['Sheet'].cell(row=(2*ii)+2, column=3,value="E0")      #テスト  E0
        wb['Sheet'].cell(row=(2*ii)+3, column=3,value="E0")      #テスト  E0
        wb['Sheet'].cell(row=(2*ii)+2, column=4,value="002845")
        wb['Sheet'].cell(row=(2*ii)+3, column=4,value="002845")
        wb['Sheet'].cell(row=(2*ii)+2, column=5,value="1016")
        wb['Sheet'].cell(row=(2*ii)+3, column=5,value="1016")
        wb['Sheet'].cell(row=(2*ii)+2, column=6,value="002845")
        wb['Sheet'].cell(row=(2*ii)+3, column=6,value="002845")
        wb['Sheet'].cell(row=(2*ii)+2, column=7,value=仕入日)
        wb['Sheet'].cell(row=(2*ii)+3, column=7,value=仕入日)
        wb['Sheet'].cell(row=(2*ii)+2, column=8,value="0")
        wb['Sheet'].cell(row=(2*ii)+3, column=8,value="0")
        wb['Sheet'].cell(row=(2*ii)+2, column=9,value="00")
        wb['Sheet'].cell(row=(2*ii)+3, column=9,value="00")
        wb['Sheet'].cell(row=(2*ii)+2, column=10,value=仕入日)
        wb['Sheet'].cell(row=(2*ii)+3, column=10,value=仕入日)
        wb['Sheet'].cell(row=(2*ii)+2, column=11,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=11,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=12,value="1")
        wb['Sheet'].cell(row=(2*ii)+3, column=12,value="1")
        wb['Sheet'].cell(row=(2*ii)+2, column=13,value="0")
        wb['Sheet'].cell(row=(2*ii)+3, column=13,value="0")
        wb['Sheet'].cell(row=(2*ii)+2, column=14,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=14,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=15,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=15,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=16,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=16,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=17,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=17,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=18,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=18,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=19,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=19,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=20,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=20,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=21,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=21,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=22,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=22,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=23,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=23,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=24,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=24,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=25,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=25,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=26,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=26,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=27,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=27,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=28,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=28,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=29,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=29,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=30,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=30,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=31,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=31,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=32,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=32,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=33,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=33,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=34,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=34,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=35,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=35,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=36,value=摘要)
        wb['Sheet'].cell(row=(2*ii)+3, column=36,value=摘要)
        wb['Sheet'].cell(row=(2*ii)+2, column=37,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=37,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=38,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=38,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=39,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=39,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=40,value=1)
        wb['Sheet'].cell(row=(2*ii)+3, column=40,value=1)
        wb['Sheet'].cell(row=(2*ii)+2, column=41,value=0)
        wb['Sheet'].cell(row=(2*ii)+3, column=41,value=1)
        wb['Sheet'].cell(row=(2*ii)+2, column=42,value=勘定科目コード)
        wb['Sheet'].cell(row=(2*ii)+3, column=42,value=2141)
        wb['Sheet'].cell(row=(2*ii)+2, column=43,value=負担部門コード)
        wb['Sheet'].cell(row=(2*ii)+3, column=43,value=1016)
        wb['Sheet'].cell(row=(2*ii)+2, column=44,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=44,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=45,value=細目コード)
        wb['Sheet'].cell(row=(2*ii)+3, column=45,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=46,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=46,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=47,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=47,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=48,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=48,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=49,value=集計拡張コード)
        wb['Sheet'].cell(row=(2*ii)+3, column=49,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=50,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=50,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=51,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=51,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=52,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=52,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=53,value=稟議)
        wb['Sheet'].cell(row=(2*ii)+3, column=53,value=稟議)
        wb['Sheet'].cell(row=(2*ii)+2, column=54,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=54,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=55,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=55,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=56,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=56,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=57,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=57,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=58,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=58,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=59,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=59,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=60,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=60,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=61,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=61,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=62,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=62,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=63,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=63,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=64,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=64,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=65,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=65,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=66,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=66,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=67,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=67,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=68).number_format = excel.styles.numbers.FORMAT_TEXT
        wb['Sheet'].cell(row=(2*ii)+2, column=68,value=取引先コード)
        wb['Sheet'].cell(row=(2*ii)+2, column=68).number_format = excel.styles.numbers.FORMAT_TEXT    
        wb['Sheet'].cell(row=(2*ii)+3, column=68,value=取引先コード)
        wb['Sheet'].cell(row=(2*ii)+2, column=69,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=69,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=70,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=70,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=71,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=71).number_format = excel.styles.numbers.FORMAT_TEXT    
        wb['Sheet'].cell(row=(2*ii)+3, column=71,value=取引先コード)
        wb['Sheet'].cell(row=(2*ii)+2, column=72,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=72,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=73,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=73,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=74,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=74,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=75,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=75,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=76,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=76,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=77,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=77,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=78,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=78,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=79,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=79,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=80,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=80,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=81,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=81,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=82,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=82,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=83,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=83,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=84,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=84,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=85,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=85,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=86,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=86,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=87,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=87,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=88,value=消費税区分コード)
        wb['Sheet'].cell(row=(2*ii)+3, column=88,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=89,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=89,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=90,value=消費税率区分)
        wb['Sheet'].cell(row=(2*ii)+3, column=90,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=91,value=tAmount)
        wb['Sheet'].cell(row=(2*ii)+3, column=91,value=tAmount)
        wb['Sheet'].cell(row=(2*ii)+2, column=92,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=92,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=93,value=消費税)
        wb['Sheet'].cell(row=(2*ii)+3, column=93,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=94,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=94,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=95,value=消費税区分)
        wb['Sheet'].cell(row=(2*ii)+3, column=95,value=0)
        wb['Sheet'].cell(row=(2*ii)+2, column=96,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=96,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=97,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=97,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=98,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=98,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=99,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=99,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=100,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=100,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=101,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=101,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=102,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=102,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=103,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=103,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=104,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=104,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=105,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=105,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=106,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=106,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=107,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=107,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=108,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=108,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=109,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=109,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=110,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=110,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=111,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=111,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=112,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=112,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=113,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=113,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=114,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=114,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=115,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=115,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=116,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=116,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=117,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=117,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=118,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=118,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=119,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=119,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=120,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=120,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=121,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=121,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=122,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=122,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=123,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=123,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=124,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=124,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=125,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=125,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=126,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=126,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=127,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=127,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=128,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=128,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=129,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=129,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=130,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=130,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=131,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=131,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=132,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=132,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=133,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=133,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=134,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=134,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=135,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=135,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=136,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=136,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=137,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=137,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=138,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=138,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=139,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=139,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=140,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=140,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=141,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=141,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=142,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=142,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=143,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=143,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=144,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=144,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=145,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=145,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=146,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=146,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=147,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=147,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=148,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=148,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=149,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=149,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=150,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=150,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=151,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=151,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=152,value=iNo)
        wb['Sheet'].cell(row=(2*ii)+3, column=152,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=153,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=153,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=154,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=154,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=155,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=155,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=156,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=156,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=157,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=157,value="P01")
        wb['Sheet'].cell(row=(2*ii)+2, column=158,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=158,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=159,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=159,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=160,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=160,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=161,value='')
        wb['Sheet'].cell(row=(2*ii)+3, column=161,value='')
        wb['Sheet'].cell(row=(2*ii)+2, column=162,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=162,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=163,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=163,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=164,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=164,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=165,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=165,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=166,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=166,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=167,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=167,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=168,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=168,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=169,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=169,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=170,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=170,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=171,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=171,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=172,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=172,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=173,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=173,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=174,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=174,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=175,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=175,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=176,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=176,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=177,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=177,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=178,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=178,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=179,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=179,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=180,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=180,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=181,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=181,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=182,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=182,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=183,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=183,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=184,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=184,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=185,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=185,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=186,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=186,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=187,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=187,value="")
        wb['Sheet'].cell(row=(2*ii)+2, column=188,value="")
        wb['Sheet'].cell(row=(2*ii)+3, column=188,value="")
        wb.save('1012_Upload_OCR.xlsx')
        ii=ii+1
        jj=jj+1
    except:
        with open("1012_ErrorList_"+str(datetime.date.today().strftime('%Y%m%d'))+".txt","a+") as f:
            f.write(str([tifs[i],company+",InvoiceNo:"+str(iNo),"Amount:"+str(tAmount)])+"\n")

# Upload_OCR用 CSVファイル書出し
#       定期支払チェックリスト.xlsx、Upload_OCR.xlsx 修正時 リラン 用

wb = excel.load_workbook('1012_Upload_OCR.xlsx')
ws = wb["Sheet"]
# ws.delete_rows(1,1)                                      # DF ヘッダー無しの場合
wb.save('Upload_OCR_kari.xlsx')

wb2 = excel.load_workbook('Upload_OCR_kari.xlsx')
ws2 = wb2.worksheets[0]
with open('1012_Upload_OCR_経理.csv', 'w', newline="", encoding='cp932') as csvfile:
    writer = csv.writer(csvfile)
    for row in ws2.rows:                                   #ヘッダー有りの場合
    #for row in ws2.iter_rows(min_row=2):                  #ヘッダー無しの場合
        writer.writerow( [cell.value for cell in row] )

#定期支払チェックリスト.xlsx 読み合わせ用 列幅印刷設定
wb3 = excel.load_workbook('1012_Teiki_CheckList.xlsx')
ws3 = wb3["Sheet"]

from openpyxl.styles.borders import Border, Side
from openpyxl.styles.colors import Color

rr = ws3.max_row
color1 = Color(rgb='000000')   #黒
side1 = Side(style='thin', color=color1)
border1 = Border(top=side1, bottom=side1, left=side1, right=side1)  # 上下左右
for rows in ws3[1:rr]:                                    
    for cell in rows:
        ws3[cell.coordinate].border = border1
ws3.column_dimensions['A'].width = 10
ws3.column_dimensions['B'].width = 10
ws3.column_dimensions['C'].width = 10
ws3.column_dimensions['D'].width = 25
ws3.column_dimensions['E'].width = 25
ws3.column_dimensions['F'].width = 10
ws3.column_dimensions['G'].width = 8
ws3.column_dimensions['H'].width = 8
ws3.delete_cols(9,3)
ws3.page_setup.orientation = 'landscape' 
wb3.save('1012_Teiki_CheckList_読み合わせ.xlsx')
#       定期支払チェックリスト.xlsx、Upload_OCR.xlsx 修正時 リラン 用

#END