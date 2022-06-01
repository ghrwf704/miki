# PC設置は、Desktop上のフォルダーに ｐｙ あるいは exe を配置して実行する。  （78行目より）
# Chrome 設定は、保存先(/Download) の、 ダウンロード前に各ファイルの保存場所を確認する を off にする

import time
import socks
import urllib3
from urllib3.exceptions import InsecureRequestWarning
urllib3.disable_warnings(InsecureRequestWarning)
from selenium import webdriver
import openpyxl as px
import datetime
import send2trash
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options


#エクセルのメール送信シートから情報を抜き出して返す
def arr_str(arr1):
    tmp=""
    for i in range(1,len(arr1)):
        tmp=tmp+arr1[i]+","
    return tmp
    
#プロキシ設定
proxies = {'https': 'swg-proxy.mikipulley.co.jp:8080',}
socks.set_default_proxy(socks.SOCKS5, "swg-proxy.mikipulley.co.jp", 8080)

#エクセルテンプレート 読み込み
wb = px.load_workbook(filename=u'secomTemplate.xlsx')

#メール送信に必要な情報を取得
ws=wb["送信先メールアドレス"]
arr=[["D"],["E"],["F"],["A"],["B"],["C"]]
for k in range(0,len(arr)):
    for i in range(5,20):
        if ws[arr[k][0]+str(i)].value!=None:
            arr[k].append(ws[arr[k][0]+str(i)].value)

#生成するエクセル名
import os
fname="セコムデータ_"+datetime.datetime.today().strftime("%Y%m%d%H%M%S")+".xlsx"

for ind in range(2,12):
    #secomTemplate.xlsxをもとにしてファイルを生成する
    ws=wb["支店一覧"]
    SITEN=ws["A"+str(ind)].value
    ID=ws["B"+str(ind)].value
    PASS=ws["C"+str(ind)].value

    #chrome起動時optionを設定
    options = Options()

    options.add_argument('--disable-gpu')


    options = Options()
    # ヘッドレスモードを有効にする（次の行をコメントアウトすると画面が表示される）。
    options.add_argument('--incognito')
    # ChromeのWebDriverオブジェクトを作成する。
    driver = webdriver.Chrome(chrome_options=options)
    driver.set_page_load_timeout(45)    
    #Javascript実行が終了するまで最大45秒間待つ
    driver.set_script_timeout(15)
    driver.get('https://secom.jp/wsr2/login/init.do')
    driver.set_script_timeout(15)
    driver.find_elements_by_tag_name('input')[1].send_keys(ID)
    driver.find_elements_by_tag_name('input')[3].send_keys(PASS)
    driver.find_element_by_id("login").click()
    driver.get('https://secom.jp/wsr2/useReport/init.do')
    driver.find_element_by_xpath("/html/body/div/form/table[2]/tbody/tr[3]/td[2]/select/option").click()
    driver.find_element_by_xpath("/html/body/div/form/table[2]/tbody/tr[1]/td[2]/input").click()
    time.sleep(1)
    driver.close()
                
    #エクセルテンプレート secomTemplate.xlsx を参照
    ws=wb.get_sheet_by_name("数式原本")
    ws["A1"].value=str(SITEN)
    ws["A2"].value=str(datetime.date.today().year)
    ws["B2"].value=str(datetime.date.today().month)
    wb.copy_worksheet(ws)
    ws=wb.get_sheet_by_name("数式原本 Copy")
    ws.title=str(SITEN) 
    j=5
    red=0
    yellow=0
    import openpyxl
    fill1 = openpyxl.styles.PatternFill(patternType='solid',fgColor='FF0000', bgColor='FF0000')
    fill2 = openpyxl.styles.PatternFill(patternType='solid',fgColor='FFFF00', bgColor='FFFF00')
    time.sleep(1)
    csvName="D:\\"+os.environ['HOMEPATH'].replace("Users","user")+"\\Downloads\\"+ID+".csv"
    with open(csvName) as f:
        while(f):
            j=j+1
            bb=f.readline().split(",")
            if bb==[""]:
                break
            if j==6:
                continue
            l=0
            for k in range(2,7):
                l=l+1
                ws.cell(row=j,column=l).value = bb[k]
                if k==4:
                    if (ws.cell(row=j,column=l).value)[0:2] in "22,23":
                        ws["c"+str(j)].fill = fill2
                        yellow=yellow+1
                    elif (ws.cell(row=j,column=l).value)[0:2] in "0:,1:,2:,3:,4:,5:":
                        ws["c"+str(j)].fill = fill1
                        red=red+1
        ws.delete_rows(idx=j, amount=1000-j+1)
        ws["c"+str(j)].value = str(red)
        ws["d"+str(j)].value = str(yellow)
    ws.freeze_panes = 'A7'
    wb.save(fname)
    ws=wb.active
    send2trash.send2trash(csvName)
wb = px.load_workbook(filename=fname)
del wb['支店一覧']
del wb['数式原本']
del wb['送信先メールアドレス']
wb.active = wb["Ａ棟"] 
wb.save(fname)
wb.close()

#メール送信
import smtplib
import ssl
from email.mime.multipart import MIMEMultipart
from email.utils import formatdate
from email.mime.application import MIMEApplication
from os.path import basename
from email.mime.text import MIMEText

def addAttach(name):
    with open(name, "rb") as f:
        part = MIMEApplication(
            f.read(),
            Name=basename(name)
        )
    part['Content-Disposition'] = 'attachment; filename="%s"' % basename(name)
    return part    

"""
(1) MIMEMultipartでメッセージを作成
"""
main_text = arr[2][-1]
charset = "utf-8"
msg = MIMEMultipart()
msg["Subject"] = arr[0][-1]
msg["From"] = arr[1][-1] 
msg["To"] = arr_str(arr[3])
msg["Cc"] = arr_str(arr[4])
msg["Bcc"] = arr_str(arr[5])
msg["Date"] = formatdate(None,True)
body = MIMEText(main_text.encode("utf-8"), 'plain', 'utf-8')
msg.attach(body)

host = "mail.securemx.jp"
nego_combo = ("ssl", 465) # ("通信方式(ssl)", port番号)
context = ssl.create_default_context()
smtpclient = smtplib.SMTP_SSL(host, nego_combo[1], timeout=10, context=context)

"""
(2) サーバーにログイン
"""
USERNAME = "ando.shinobu@mikipulley.co.jp"
PASSWORD = "h5yZ3q64GF"                                        # パスワード指定
smtpclient.login(USERNAME, PASSWORD)

"""
(3) 添付ファイル追加
"""
msg.attach(addAttach(fname))

"""
(4) メールを送信する
"""
smtpclient.send_message(msg)
smtpclient.quit()

import shutil
shutil.copy(fname,"\\\\Serg001\\dat12\\●総務\\●全社公開用\\RPA\\セコム入退室データ")
send2trash.send2trash(fname)
print('処理が完了しました。')
print('プロンプト画面を閉じて、終了して下さい。')