import subprocess
import openpyxl
import re
import datetime
import os
import csv
import requests
import time
import socks
import json
from PIL import Image

def regmatch(src,pat):
	search_result = re.search(pat, src)
	if search_result:
	  return search_result.group()
	else:
	  return 0
    
def readImg(temp):
    #アップロード用のphpファイルに接続
    url = 'http://mikipulley.sakura.ne.jp/microsoft/azuru/ocr/img/upload.php'
    file = {'upfile': open(os.getcwd()+"\\undone\\"+temp, 'rb')}
    res = requests.post(url, files = file, proxies = proxies)#画像ファイルなどをアップするのに必要
    
    #ocr用のコマンドを作成
    url = "https://x8g7u5g3e0.apigw.ntruss.com/custom/v1/81/507d6851a2bcac504757c058f106a6d1d04216ea394211eb265e7098dc55afef/infer"
    payload = "{\r\n    \"images\": [\r\n      {\r\n        \"format\": \"png\",\r\n        \"name\": \"All\",\r\n        \"data\": null,\r\n        \"url\": \""+ 'http://mikipulley.sakura.ne.jp/microsoft/azuru/ocr/img/files/'+temp+"\"\r\n      }\r\n    ],\r\n    \"lang\": \"ja\",\r\n    \"requestId\": \"string\",\r\n    \"resultType\": \"string\",\r\n    \"timestamp\": 1582166299,\r\n    \"version\": \"V1\"\r\n}"
    headers = {
            'Content-Type': 'application/json',
            'X-OCR-SECRET': 'TlVJcVBraWFEVWFLUlN1b1hmT3RSYWlvbVRUY0ZwTnY='
            }
    try:
        #コマンド送信
        response = requests.request("POST", url, headers=headers, data = payload, proxies = proxies)
        #戻り値にエラーがあった場合はすべての値を空白にして返す（合致するテンプレートがない場合もエラーになります）
        with open(temp.replace(".png",".txt"), 'w') as f:
            json.dump(response.json(), f, ensure_ascii=False)
        company=response.json()["images"][0]["matchedTemplate"]["name"]
        iNo=response.json()["images"][0]["fields"][0]["inferText"]
        tAmount=response.json()["images"][0]["fields"][1]["inferText"]
    except:
        company=""
        iNo=""
        tAmount=""
    return company.replace("MPHK_",""),regmatch(iNo,"[０-９0-9a-zA-Z\-\(\)]+"),regmatch(tAmount,"\d{1,3}(,\d{3})*(\.\d+)?").replace("・",".")

def getCompanyInfo(company):
    temp=["仮"]
    with open('regist_data.csv') as f:
        while temp[0]!="":
            temp=f.readline().split(",")
            if company in temp:
                if len(temp)!=4:
                    break
                return temp[0:4]
        return ["","","",""]

#プロキシ設定
proxies = {
  'https': 'swg-proxy.mikipulley.co.jp:8080',
}
socks.set_default_proxy(socks.SOCKS5, "swg-proxy.mikipulley.co.jp", 8080)

files = os.listdir(os.getcwd()+"\\undone")
for k in files:
    if ".tif" in k:
        pil_img = Image.open(os.getcwd()+"\\undone\\"+k,'r')
        pil_img.save(os.getcwd()+"\\undone\\"+k.replace(".tif",".png"), 'png')
#エクセル起動
wb = openpyxl.load_workbook('Upload_Voucher_Template.xlsx')
ws = wb['Voucher']
ws.delete_rows(idx=2, amount=1000)
ws = wb['Document Transaction']
ws.delete_rows(idx=2, amount=1000)
wb.save('Upload_Voucher_Template.xlsx')


#現在の年月日情報を取得
time_now = datetime.datetime.now()
year = time_now.strftime("%Y")
month = time_now.strftime("%m")
day = time_now.strftime("%d")

#対象のファイル名を書き出す（ｓｔｒ１にカンマ区切りで格納しておく)
subprocess.call("dir /B "+os.getcwd()+"\\undone\\*.png>"+os.getcwd()+"\\file.txt",shell=True)
str1=""
t=open('file.txt')
f=t.readline()
while f:
	if ""==f:
		continue
	if "tmp.png"==f:
		continue
	str1=str1+f+","
	f=t.readline()
t.close
str1=str1.replace("\n","")
tifs=str1.split(",")
ocr_output=""
ii=0
company="仮"
for i in range(0, len(tifs)):
    [company,iNo,tAmount]=readImg(tifs[i])
    print([company,iNo,tAmount,tifs[i]])
    #excel操作
    wb['Voucher'].cell(row=ii+2, column=2,value=ii+1)
    wb['Voucher'].cell(row=ii+2, column=3,value=str(day+"/"+month+"/"+year))
    wb['Voucher'].cell(row=ii+2, column=4,value="JV")
    wb['Document Transaction'].cell(row=(2*ii)+2, column=2,value=ii+1)
    wb['Document Transaction'].cell(row=(2*ii)+3, column=2,value=ii+1)
    wb['Document Transaction'].cell(row=(2*ii)+2, column=3,value=1)
    wb['Document Transaction'].cell(row=(2*ii)+3, column=3,value=2)
    wb['Document Transaction'].cell(row=(2*ii)+2, column=6,value="HKD")
    wb['Document Transaction'].cell(row=(2*ii)+3, column=6,value="HKD")
    #companyのデータをresist.csvから取得
    [symbol,name,code1,code2]=getCompanyInfo(company)    
    wb['Document Transaction'].cell(row=(2*ii)+2, column=4,value=code1)
    wb['Document Transaction'].cell(row=(2*ii)+3, column=4,value=code2)
    wb['Document Transaction'].cell(row=(2*ii)+2, column=9,value=name)
    wb['Document Transaction'].cell(row=(2*ii)+3, column=9,value=name)
    wb['Document Transaction'].cell(row=(2*ii)+2, column=10,value=tAmount)
    wb['Document Transaction'].cell(row=(2*ii)+2, column=11,value=tAmount)
    wb['Document Transaction'].cell(row=(2*ii)+3, column=12,value=tAmount)
    wb['Document Transaction'].cell(row=(2*ii)+3, column=13,value=tAmount)
    wb['Document Transaction'].cell(row=(2*ii)+3, column=14,value=tifs[i])
    wb['Document Transaction'].cell(row=(2*ii)+2, column=35,value=iNo)
    wb['Document Transaction'].cell(row=(2*ii)+3, column=35,value=iNo)
    if((code1==51270) or (code2==51270)):
        wb['Document Transaction'].cell(row=2*(i),column=16,value="TRD")
    wb.save('./Upload_Voucher_Template.xlsx')
    ii=ii+1
    