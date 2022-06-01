import subprocess
import openpyxl
import re
import datetime
import os
import csv
import cv2
import requests
import time
import json
from PIL import Image
import pdf2image
import PyPDF2
import shutil
import numpy as np
# import socks   #proxy設定時

####################使い方###########################
# 本スクリプト（または実行）ファイルと自動転記用のエクセル、設定ファイル 
# regist.csv を同一フォルダにおく。
# 読取りファイル(tif png pdf 形式)は、「undone」フォルダーに格納、
# 処理tempデータも同フォルダーに保存される
# 処理エラーは、実行Fに、 ErrorList.txt で保存される。
###################################################

#imgのバイナリからpngファイルを生成
#cvの関数でなく、自分で定義（バグがあって2byte表示は文字化けする）
def imwrite(filename, img, params=None):
    try:
        ext = os.path.splitext(filename)[1]
        result, n = cv2.imencode(ext, img, params)
        if result:
            with open(filename, mode='w+b') as f:
                n.tofile(f)
            return True
        else:
            return False
    except Exception as e:
        print(e)
        return False

#複数のPDFファイルでもページ分割するように修正
def split_pdf_pages(src_path):
    dst_basepath=src_path.replace(".pdf","")
    src_pdf = PyPDF2.PdfFileReader(src_path)
    for i in range(src_pdf.numPages):
        if i==src_pdf.numPages:
            return -1
        dst_pdf = PyPDF2.PdfFileWriter()
        dst_pdf.addPage(src_pdf.getPage(i))
        if os.path.isfile('{}_{}.pdf'.format(dst_basepath, i)):
            return -1
        with open('{}_{}.pdf'.format(dst_basepath, i), 'wb') as f:
            dst_pdf.write(f)
            
#正規表現（請求書No、合計金額の取得に使用）
def regmatch(src,pat):
	search_result = re.search(pat, src)
	if search_result:
	  return search_result.group()
	else:
	  return 0
    
def readImg(temp):
    #アップロード用のphpファイルに接続
    url = 'http://mikipulley.sakura.ne.jp/microsoft/azuru/ocr/img/upload.php'
    file = {'upfile': open(os.getcwd()+"\\undone\\"+temp, 'rb')}
    res = requests.post(url, files = file)#画像ファイルなどをアップするのに必要
    # res = requests.post(url, files = file, proxies = proxies) # proxy設定時
    
    #ocr用のコマンドを作成
    url = "https://x8g7u5g3e0.apigw.ntruss.com/custom/v1/142/6a6b3f6f073a097b717ebc912bc844c01612087482d602f094292fef8e3ec9f1/infer"
    payload = "{\r\n    \"images\": [\r\n      {\r\n        \"format\": \"png\",\r\n        \"name\": \"All\",\r\n        \"data\": null,\r\n        \"url\": \""+ 'http://mikipulley.sakura.ne.jp/microsoft/azuru/ocr/img/files/'+temp+"\"\r\n      }\r\n    ],\r\n    \"lang\": \"ja\",\r\n    \"requestId\": \"string\",\r\n    \"resultType\": \"string\",\r\n    \"timestamp\": 1582166299,\r\n    \"version\": \"V1\"\r\n}"
    headers = {
            'Content-Type': 'application/json',
            'X-OCR-SECRET': 'WHVJREREWHZ4TnRrQVJURWxyRlNsb3RGa3VaWnBJVlg='
            }
    try:
        #リクエスト回数節約のため、ocrの結果を保存しておく
        response = requests.request("POST", url, headers=headers, data = payload)
        # response = requests.request("POST", url, headers=headers, data = payload, proxies = proxies)  #proxy設定時
        #戻り値にエラーがあった場合はすべての値を空白にして返す（合致するテンプレートがない場合もエラーになります）
        with open("undone\\"+temp.replace(".png",".txt"), 'w') as f:
            json.dump(response.json(), f, ensure_ascii=False)
        time.sleep(3)
        #保存しておいたテキストからjsonとして読み込む
        json_open = open('undone\\'+tifs[i].replace(".png",".txt"), 'r')
        response = json.load(json_open)
        company=response["images"][0]["matchedTemplate"]["name"]
        iNo=response["images"][0]["fields"][0]["inferText"]
        tAmount=response["images"][0]["fields"][1]["inferText"]
        iNoConf=response["images"][0]["fields"][0]["inferConfidence"]
        tAmountConf=response["images"][0]["fields"][1]["inferConfidence"]
    except:
        company=""
        iNo=""
        tAmount=""
        iNoConf=0
        tAmountConf=0
    return company.replace("MPHK_",""),iNo,tAmount,iNoConf,tAmountConf

def getCompanyInfo(company):
    temp=["仮"]
    with open('regist_data.csv') as f:
        rows=csv.reader(f)
        for temp in rows:
            if ""==company:
                break
            if company in str(temp):
                return temp[0:6]
            if len(temp)!=6:
                break
    return ["","","",0,100,""]

"""
#プロキシ設定
proxies = {
  'https': 'swg-proxy.mikipulley.co.jp:8080',
}
socks.set_default_proxy(socks.SOCKS5, "swg-proxy.mikipulley.co.jp", 8080)
"""

#カレントディレクトリー変更
path=os.getcwd()+'\\undone'
files = os.listdir(path)
files_file = [f for f in files if os.path.isfile(os.path.join(path, f))]

#PDFをいページずつ分けるのをやめる
#for i in range(len(files_file)):
#    if ".pdf" in files_file[i]:
#        # .spyder-py3\Data フォルダーの、 結合されている merge.pdf から  split_0,split_1,...へ 分割
#        split_pdf_pages(os.getcwd()+'\\undone\\'+files_file[i])   
#        #os.remove("undone\\"+files_file[i])

os.getcwd()
q = os.chdir("undone")
filepath = os.listdir(q)
#ファイル名から" "空白と"#"を置換
for file in filepath:
    os.rename(file,file.replace(" ","_").replace("#","_"))
    print(file)
#os.system("del /q \""+os.getcwd().replace("\\Data","\\CVImage")+"\\*.*\"")
os.chdir("../")


#PDFをⅠページづつに分解してそれぞれPNGに変換する
for file in filepath:
    basename = os.path.basename(file)
    if ".pdf" in basename:
        pdfimages = pdf2image.convert_from_path("undone\\"+file)
        cvimage = np.asarray(pdfimages[0])
        cvimage = cv2.cvtColor(cvimage, cv2.COLOR_BGR2GRAY)
        #cvimage = cv2.resize(cvimage, (1920, 1050))
        temp=str(basename).replace('.pdf','.png')
        imwrite('undone\\'+temp, cvimage)
        #send2trash.send2trash("undone\\"+basename+".pdf")

files = os.listdir(os.getcwd()+"\\undone")
for k in files:
    if ".tif" in k:
        pil_img = Image.open(os.getcwd()+"\\undone\\"+k,'r')    
        pil_img.save(os.getcwd()+"\\undone\\"+k.replace(".tif",".png"), 'png')
#エクセル起動
wb = openpyxl.load_workbook('Upload_Voucher_Template.xlsx')
ws = wb['Voucher']
ws.delete_rows(idx=2, amount=1000)
ws = wb['Document Transaction']
ws.delete_rows(idx=2, amount=1000)
wb.save('./Upload_Voucher_Template.xlsx')


#現在の年月日情報を取得
time_now = datetime.datetime.now()
year = time_now.strftime("%Y")
month = time_now.strftime("%m")
day = time_now.strftime("%d")

#対象のファイル名を書き出す（ｓｔｒ１にカンマ区切りで格納しておく)
subprocess.call("dir /B "+os.getcwd()+"\\undone\\*.png>"+os.getcwd()+"\\tempfile.txt",shell=True)
str1=""
t=open('tempfile.txt')
f=t.readline()
while f:
	if ""==f:
		continue
	if "tmp.png"==f:
		continue
	str1=str1+f+","
	f=t.readline()
t.close()
str1=str1.replace("\n","")
#ファイル名が格納された配列作成（tifs配列にはファイル名のみが入っている）
tifs=str1.split(",")
os.remove("tempfile.txt")
ocr_output=""
ii=0
company="仮"

for i in range(0, len(tifs)):
    if tifs[i]=="":
        break
    #PDFを画像に変換、テンプレートが一致したら、INVOICE番号と、合計金額を取得
    [company,iNo,tAmount,conf1,conf2]=readImg(tifs[i]) 
    #companyのデータをresist.csvから取得
    #csvのA列のデータ==companyになってるのを探して行情報を返す
    [symbol,name,code1,code2,ritu,reg]=getCompanyInfo(company)    
    tAmount=tAmount.split("\n")[-1]
    #tAmountに格納されている文字列データのうち、最下段文字列の1-0のみmを取得
    #改行以降を削除したうえで、数値のみを連結させる
    match = re.findall(r'[0-9]+', re.sub('\n.+',"",tAmount))
    s=""
    #連結（飛び飛びでもかいつまんで連結させる）※1,2,er45,t77→124577→csvファイルの一番右の列がで割る（124577 or 1245.77）
    for m in match:
	    s=s+str(m)
    if s=="":
        s=0
    tAmount=int(s)
    print(str([tifs[i],"RegistName：MPHK_"+company+",InvoiceNo:"+str(iNo),"Amount:"+str(tAmount),code1,code2,name])+"\n")
    #iNoを正規表現で抽出する 空白削除
    iNo=regmatch(iNo,reg).replace(" ","").replace("　","").replace("\t","")
    if iNo==0 or iNo=="0":
        iNo=""
    tAmount=tAmount/int(ritu)  #金額小数処理
    if company=="" or code1==0 or code2==0 or tAmount==0:#(conf1)<0.8 or (conf2)<0.8
        with open("ErrorList.txt","a+") as f:
            f.write(str([tifs[i],"RegistName：MPHK_"+company+",InvoiceNo:"+str(iNo),"Amount:"+str(tAmount),code1,code2,name])+"\n")
        continue
    #excel操作
    wb['Voucher'].cell(row=ii+2, column=2,value=ii+1)
    wb['Voucher'].cell(row=ii+2, column=3,value=str(day+"/"+month+"/"+year))
    wb['Voucher'].cell(row=ii+2, column=4,value="JV")
    wb['Document Transaction'].cell(row=(2*ii)+2, column=2,value=ii+1)
    wb['Document Transaction'].cell(row=(2*ii)+3, column=2,value=ii+1)
    wb['Document Transaction'].cell(row=(2*ii)+2, column=3,value=1)
    wb['Document Transaction'].cell(row=(2*ii)+3, column=3,value=2)
    wb['Document Transaction'].cell(row=(2*ii)+2, column=6,value="HKD")
    wb['Document Transaction'].cell(row=(2*ii)+3, column=6,value="HKD")
    if code1!="XX":
        code1=int(code1)
        code2=int(code2)
    wb['Document Transaction'].cell(row=(2*ii)+2, column=4,value=(code1))
    wb['Document Transaction'].cell(row=(2*ii)+3, column=4,value=(code2))
    wb['Document Transaction'].cell(row=(2*ii)+2, column=9,value=name)
    wb['Document Transaction'].cell(row=(2*ii)+3, column=9,value=name)
    wb['Document Transaction'].cell(row=(2*ii)+2, column=10,value=float(tAmount))
    wb['Document Transaction'].cell(row=(2*ii)+2, column=11,value=float(tAmount))
    wb['Document Transaction'].cell(row=(2*ii)+3, column=12,value=float(tAmount))
    wb['Document Transaction'].cell(row=(2*ii)+3, column=13,value=float(tAmount))
    # wb['Document Transaction'].cell(row=(2*ii)+3, column=14,value=tifs[i]) #読取りpngファイル名書出し
    wb['Document Transaction'].cell(row=(2*ii)+2, column=35,value=iNo)
    wb['Document Transaction'].cell(row=(2*ii)+3, column=35,value=iNo)
    if((code1==51270) or (code2==51270)):
        wb['Document Transaction'].cell(row=(2*ii)+2,column=16,value="TRD")
    wb.save('./Upload_Voucher_Template.xlsx')
    ii=ii+1
    